import openpyxl
from openpyxl.styles import Font, Alignment
import random

# 科目-模块映射
SUBJECTS = {
    "言语理解与表达": ["片段阅读", "逻辑填空", "语句表达"],
    "数量关系": ["数学运算", "数字推理"],
    "判断推理": ["图形推理", "定义判断", "类比推理", "逻辑判断"],
    "资料分析": ["文字材料", "表格材料", "图形材料"],
    "常识判断": ["政治常识", "法律常识", "经济常识"],
}

# 知识点映射
KNOWLEDGE_POINTS = {
    "片段阅读": ["主旨概括", "意图判断", "细节理解", "标题选择"],
    "逻辑填空": ["实词辨析", "成语辨析", "语境分析"],
    "语句表达": ["语句排序", "语句衔接", "下文推断"],
    "数学运算": ["行程问题", "工程问题", "排列组合", "概率问题", "利润问题", "几何问题"],
    "数字推理": ["等差数列", "等比数列", "递推数列", "幂次数列"],
    "图形推理": ["位置规律", "样式规律", "数量规律", "属性规律", "空间重构"],
    "定义判断": ["单定义", "多定义"],
    "类比推理": ["逻辑关系", "言语关系", "经验常识", "理论常识"],
    "逻辑判断": ["翻译推理", "真假推理", "分析推理", "加强论证", "削弱论证"],
    "文字材料": ["增长率", "增长量", "比重", "平均数", "倍数"],
    "表格材料": ["数据查找", "数据比较", "数据计算"],
    "图形材料": ["柱状图", "折线图", "饼状图"],
    "政治常识": ["时政热点", "党史党建", "重要会议"],
    "法律常识": ["宪法", "民法", "刑法", "行政法"],
    "经济常识": ["宏观经济", "微观经济", "市场经济"],
}

# 题目模板库
TEMPLATES = {
    "片段阅读": [
        ("随着社会的不断发展，{topic}已经成为人们关注的焦点。然而，{topic}在推动社会进步的同时，也带来了一系列新的挑战。因此，我们需要在发展的过程中不断探索和完善相关机制。\n\n这段文字的主旨是：",
         ["{topic}受到人们关注", "{topic}带来新挑战", "需要完善{topic}相关机制", "社会发展不断进步"], "C"),
        ("{topic}是现代社会的重要组成部分。研究表明，{topic}能够有效提升工作效率，改善生活质量。但也有人认为，过度依赖{topic}可能导致一些问题。\n\n作者对{topic}的态度是：",
         ["完全支持", "完全反对", "客观理性", "漠不关心"], "C"),
    ],
    "逻辑填空": [
        ("在_____的{topic}背景下，只有坚持_____，才能不断取得新的突破。那些_____、停滞不前的人，终将被时代淘汰。\n\n依次填入横线处最恰当的一项是：",
         ["激烈 创新 固步自封", "严峻 改革 因循守旧", "复杂 突破 墨守成规", "艰难 发展 按部就班"], "B"),
    ],
    "数学运算": [
        ("一项工程，甲单独完成需要{days1}天，乙单独完成需要{days2}天。两人合作需要多少天完成？",
         ["{a}天", "{b}天", "{c}天", "{d}天"], None),
        ("某商品原价{price}元，先涨价{pct1}%，再降价{pct2}%，现价是多少元？",
         ["{a}元", "{b}元", "{c}元", "{d}元"], None),
        ("从{total}人中选{select}人参加会议，有多少种不同的选法？",
         ["{a}种", "{b}种", "{c}种", "{d}种"], None),
    ],
    "图形推理": [
        ("观察下列图形序列的规律，选择问号处应填入的图形。\n图形序列：圆→三角形→正方形→？",
         ["五边形", "六边形", "菱形", "梯形"], "A"),
        ("左边给定的是纸盒的外表面，下列哪一项能由它折叠而成？\n（请根据空间想象能力判断）",
         ["A", "B", "C", "D"], "B"),
    ],
    "定义判断": [
        ("{topic}是指{definition}。\n\n根据上述定义，下列属于{topic}的是：",
         ["选项A描述", "选项B描述", "选项C描述", "选项D描述"], "C"),
    ],
    "逻辑判断": [
        ("有研究表明，{topic}与工作效率呈正相关。因此，有人认为提高{topic}就能提升工作效率。\n\n以下哪项如果为真，最能削弱上述论证？",
         ["{topic}高的员工学历也更高", "工作效率还受其他因素影响", "{topic}可以通过培训提高", "工作效率与{topic}无关"], "A"),
    ],
    "资料分析": [
        ("某地区2023年GDP为{a}亿元，2024年GDP为{b}亿元。2024年GDP的增长率为：",
         ["{opt_a}%", "{opt_b}%", "{opt_c}%", "{opt_d}%"], None),
        ("某公司2024年第一季度销售额为{a}万元，第二季度为{b}万元，第三季度为{c}万元。第二季度的环比增长率是：",
         ["{opt_a}%", "{opt_b}%", "{opt_c}%", "{opt_d}%"], None),
    ],
    "常识判断": [
        ("下列关于{topic}的说法，正确的是：",
         ["说法A", "说法B", "说法C", "说法D"], "D"),
        ("党的二十大报告指出，{topic}是中国式现代化的重要特征。\n\n以下哪项最符合报告精神？",
         ["选项A", "选项B", "选项C", "选项D"], "C"),
    ],
}

# 为没有专门模板的模块使用通用模板
def get_template(module):
    if module in TEMPLATES:
        return random.choice(TEMPLATES[module])
    # 通用模板
    return (
        f"这是一道关于{module}的测试题，请根据所学知识选择正确答案。\n\n题目描述：某情境下需要运用{module}相关知识进行分析判断。",
        ["选项A", "选项B", "选项C", "选项D"],
        random.choice(["A", "B", "C", "D"])
    )

def generate_question(idx):
    subject = random.choice(list(SUBJECTS.keys()))
    module = random.choice(SUBJECTS[subject])
    knowledge = random.choice(KNOWLEDGE_POINTS.get(module, ["综合"]))

    template = get_template(module)
    content_template, opts_template, answer_template = template

    # 填充变量
    topic = random.choice(["科技创新", "绿色发展", "数字经济", "乡村振兴", "社会治理", "教育改革", "医疗健康", "文化传承"])
    days1 = random.randint(8, 20)
    days2 = random.randint(10, 30)
    price = random.randint(100, 500)
    pct1 = random.randint(10, 30)
    pct2 = random.randint(5, 20)
    total = random.randint(8, 15)
    select = random.randint(2, 5)

    content = content_template.format(
        topic=topic, days1=days1, days2=days2, price=price, pct1=pct1, pct2=pct2, total=total, select=select,
        a=random.randint(5, 12), b=random.randint(6, 15), c=random.randint(7, 18), d=random.randint(8, 20),
        definition="某种特定的行为或现象",
    )

    # 计算数学题的答案
    if "两人合作需要多少天" in content:
        from math import gcd
        lcm = days1 * days2 // gcd(days1, days2)
        ans_val = round(lcm / (days1 + days2) * days1 * days2 / lcm, 1)
        # 简化：合作天数 = 1 / (1/days1 + 1/days2)
        from fractions import Fraction
        f = Fraction(1, 1) / (Fraction(1, days1) + Fraction(1, days2))
        ans_days = f.numerator / f.denominator
        # 找一个接近的整数答案
        ans_int = round(ans_days)
        opts = [f"{ans_int}天", f"{ans_int+1}天", f"{ans_int+2}天", f"{ans_int-1}天"]
        answer = "A"
    elif "增长率" in content and "GDP" in content:
        a_val = random.randint(4000, 8000)
        b_val = int(a_val * (1 + random.randint(5, 15) / 100))
        rate = round((b_val - a_val) / a_val * 100, 1)
        content = f"某地区2023年GDP为{a_val}亿元，2024年GDP为{b_val}亿元。2024年GDP的增长率为："
        opts = [f"{rate}%", f"{rate+2}%", f"{rate-2}%", f"{rate+5}%"]
        answer = "A"
    elif "环比增长率" in content:
        a_val = random.randint(200, 500)
        growth = random.randint(5, 20)
        b_val = int(a_val * (1 + growth / 100))
        c_val = int(b_val * (1 + random.randint(-10, 20) / 100))
        rate = round((b_val - a_val) / a_val * 100, 1)
        content = f"某公司第一季度销售额为{a_val}万元，第二季度为{b_val}万元，第三季度为{c_val}万元。第二季度的环比增长率是："
        opts = [f"{rate}%", f"{rate+3}%", f"{rate-3}%", f"{rate+5}%"]
        answer = "A"
    else:
        opts = [f"选项{chr(65+i)}" for i in range(4)]
        if answer_template:
            answer = answer_template
        else:
            answer = random.choice(["A", "B", "C", "D"])

    difficulty = random.randint(1, 5)
    estimated = random.choice([45, 60, 90, 120])
    frequency = random.choice(["HIGH", "MEDIUM", "LOW"])

    return [
        subject, module, knowledge, "SINGLE", difficulty,
        content, opts[0], opts[1], opts[2], opts[3],
        answer, f"本题考查{module}中的{knowledge}，需要运用相关知识进行分析。", "模拟题库", frequency, estimated,
    ]

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题目"

    headers = ["科目", "模块", "知识点", "题型", "难度", "题干", "选项A", "选项B", "选项C", "选项D", "答案", "解析", "来源", "考频", "预估时间(秒)"]
    ws.append(headers)

    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 生成100道题
    for i in range(100):
        row = generate_question(i)
        ws.append(row)

    # 调整列宽
    col_widths = [12, 12, 10, 8, 6, 50, 20, 20, 20, 20, 6, 40, 10, 8, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output_path = "D:/CODE/zhikao-app/test_questions.xlsx"
    wb.save(output_path)
    print(f"已生成 {output_path}，共 100 道题")

if __name__ == "__main__":
    main()
